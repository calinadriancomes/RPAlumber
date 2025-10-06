# -*- coding: utf-8 -*-
"""
lumber5.py — ETL pentru "Cerere Unică" -> Plan de fertilizare
- Interfață Tkinter (similară cu ce aveai)
- Fără xlwings; doar pdfplumber, pandas, openpyxl
- Salvează în același folder cu PDF-ul selectat (fără concatenări greșite de cale)

"""

## Prerequisites packages

"""
pip install pdfplumber
pip install pandas
pip install openpyxl

"""

import os
import re
import pdfplumber
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import filedialog, messagebox
from datetime import datetime

# Regex-uri robuste
NUMERIC_PARCEL = re.compile(r"([1-9][0-9]{0,3}|1000)[a-i]$", re.IGNORECASE)
KEY_TOKEN      = re.compile(r"(?!(ZN|ZIE))\b(TA|TAn|CP|PP|CPn|PPn|PPi|TAi|TS)\b", re.IGNORECASE)
DECIMAL        = re.compile(r"^\d+(?:[.,]\d+)?$")
COD_PACHET     = re.compile(r"^(nu|[ABC]\d{2})$", re.IGNORECASE)

def _norm(v):
    if v is None: return ""
    s = str(v).strip()
    return re.sub(r"\s+", " ", s)

def _to_float(s):
    try:
        return float(str(s).replace(",", "."))
    except Exception:
        return None

def _extract_tables_to_excel(pdf_path, out_xlsx):
    """Extrage tabele simple din PDF (pdfplumber) -> output_tables.xlsx"""
    with pdfplumber.open(pdf_path) as pdf, pd.ExcelWriter(out_xlsx) as writer:
        found_any = False
        for i, page in enumerate(pdf.pages):
            table = page.extract_table()
            if not table:
                continue
            df = pd.DataFrame(table)
            # prepunem un rând gol
            empty = pd.DataFrame([[None] * len(df.columns)], columns=df.columns)
            df = pd.concat([empty, df], ignore_index=True)
            df.to_excel(writer, sheet_name=f"Table{i+1}", index=False, header=False)
            found_any = True
        if not found_any:
            # tot creăm un fișier valid
            pd.DataFrame([[]]).to_excel(writer, sheet_name="Table1", index=False, header=False)

def _scan_tables_for_rows(in_xlsx):
    """
    Returnează o listă de rânduri parsate cu chei: {'categorie','cultura','pachet','siruta','area'}.
    Parcurge foile Table* și nu se bazează pe poziții fixe de coloane.
    """
    rows = []
    xlsx = pd.ExcelFile(in_xlsx)
    for sheet in xlsx.sheet_names:
        try:
            df = xlsx.parse(sheet, header=None)
        except Exception:
            continue
        if df is None or df.empty:
            continue
        for _, r in df.iterrows():
            vals = [_norm(v) for v in r.values.tolist()]
            if all(v == "" for v in vals):
                continue
            # găsește parcele tip "12a"
            parcel_idx = -1
            for idx, v in enumerate(vals):
                if NUMERIC_PARCEL.search(v):
                    parcel_idx = idx
                    break
            if parcel_idx == -1:
                continue
            # găsește categoria (TA/PP/CP/...)
            categorie = None
            for idx, v in enumerate(vals):
                if KEY_TOKEN.fullmatch(v):
                    categorie = v.upper()
                    break
            if categorie is None:
                continue
            # aria = ultimul număr din rând
            area = None
            area_idx = -1
            for idx in range(len(vals)-1, -1, -1):
                if DECIMAL.fullmatch(vals[idx]):
                    f = _to_float(vals[idx])
                    if f is not None:
                        area = f
                        area_idx = idx
                        break
            if area is None:
                continue
            # cultura: imediat la dreapta parcelei sau cel mai "alfabetic" text între parcelă și arie
            cultura = None
            if parcel_idx + 1 < len(vals) and vals[parcel_idx+1]:
                cultura = vals[parcel_idx+1]
            if not cultura:
                span = vals[parcel_idx+1: area_idx if area_idx > parcel_idx else len(vals)]
                alpha_like = [x for x in span if re.search(r"[A-Za-zĂÂÎȘȚăâîșț]", x)]
                if alpha_like:
                    cultura = max(alpha_like, key=len)
            if not cultura:
                cand = [v for v in vals if v and not KEY_TOKEN.fullmatch(v) and not DECIMAL.fullmatch(v)]
                cultura = cand[0] if cand else "NECUNOSCUT"

            # opționale
            pachet = None
            siruta = None
            for v in vals:
                if COD_PACHET.fullmatch(v):
                    pachet = v.upper()
                if re.fullmatch(r"\d{6}", v):
                    siruta = v

            rows.append({
                "categorie": categorie,
                "cultura": cultura,
                "pachet": pachet,
                "siruta": siruta,
                "area": round(area, 2)
            })
    return rows

def _aggregate(rows):
    cultura_area = {}
    categorie_area = {}
    for row in rows:
        a = float(row["area"] or 0.0)
        cultura = row.get("cultura")
        categorie = row.get("categorie")
        if cultura:
            cultura_area[cultura] = round(cultura_area.get(cultura, 0.0) + a, 2)
        if categorie:
            categorie_area[categorie] = round(categorie_area.get(categorie, 0.0) + a, 2)
    return cultura_area, categorie_area

def _write_plan(plan_path, cultura_area, categorie_area, source_pdf, an_text=None):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Plan_culturi"
    ws1.append(["Cultura", "Suprafata_ha"])
    for k, v in sorted(cultura_area.items()):
        ws1.append([k, v])
    for c in range(1, 3):
        ws1.column_dimensions[get_column_letter(c)].width = 28

    ws2 = wb.create_sheet("Plan_categorii")
    ws2.append(["Categorie", "Suprafata_ha"])
    for k, v in sorted(categorie_area.items()):
        ws2.append([k, v])
    for c in range(1, 3):
        ws2.column_dimensions[get_column_letter(c)].width = 28

    meta = wb.create_sheet("Meta")
    meta.append(["Generat la", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    meta.append(["Sursa PDF", source_pdf])
    if an_text:
        meta.append(["An (din interfață)", an_text])
    if not cultura_area and not categorie_area:
        meta.append(["Observatie", "Parserul nu a putut extrage rânduri valide; verificați formatul PDF."])

    wb.save(plan_path)

def process_file(pdf_path, an_text=None):
    """
    Pipeline complet:
    - extrage tabele -> output_tables.xlsx (în același folder cu PDF-ul)
    - parsează rândurile
    - agregă pe cultură și pe categorie
    - scrie "Plan_de_fertilizare.xlsx"
    """
    pdf_path = os.path.abspath(pdf_path)
    out_dir = os.path.dirname(pdf_path)
    out_tables = os.path.join(out_dir, "output_tables.xlsx")
    plan_path = os.path.join(out_dir, "Plan_de_fertilizare.xlsx")

    _extract_tables_to_excel(pdf_path, out_tables)
    rows = _scan_tables_for_rows(out_tables)
    cultura_area, categorie_area = _aggregate(rows)
    _write_plan(plan_path, cultura_area, categorie_area, pdf_path, an_text=an_text)
    return out_tables, plan_path, (len(rows), len(cultura_area), len(categorie_area))

# ---------------------- Interfața Tkinter ----------------------

def run_gui():
    root = tk.Tk()
    root.title("CERERE UNICĂ DE PLATĂ - ETL")
    root.geometry("520x220")

    # Linie 1: An
    frm_an = tk.Frame(root)
    frm_an.pack(padx=10, pady=(12,4), fill="x")
    tk.Label(frm_an, text="CERERE UNICĂ DE PLATĂ din anul").pack(side="left")
    ent_an = tk.Entry(frm_an, width=8)
    ent_an.insert(0, "2024")
    ent_an.pack(side="left", padx=(6,0))

    # Linie 2: Selectare PDF
    frm_pdf = tk.Frame(root)
    frm_pdf.pack(padx=10, pady=6, fill="x")
    tk.Label(frm_pdf, text="Selectați un document PDF IPA-Online:", fg="red").pack(anchor="w")
    path_var = tk.StringVar(value="(Neselectat)")
    lbl_path = tk.Label(frm_pdf, textvariable=path_var, anchor="w", justify="left")
    lbl_path.pack(fill="x")

    state = {"pdf": None}

    def choose_pdf():
        fn = filedialog.askopenfilename(
            title="Selectați PDF-ul IPA-Online",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
        )
        if fn:
            state["pdf"] = fn
            path_var.set(fn)

    def do_process():
        if not state["pdf"]:
            messagebox.showwarning("Lipsă fișier", "Selectați mai întâi un PDF.")
            return
        try:
            out_tables, plan, stats = process_file(state["pdf"], an_text=ent_an.get().strip())
            messagebox.showinfo(
                "Succes",
                "Fișiere create în același folder cu PDF-ul:\n"
                f"- {os.path.basename(out_tables)}\n"
                f"- {os.path.basename(plan)}\n\n"
                f"Statistici: rânduri parse: {stats[0]}, culturi: {stats[1]}, categorii: {stats[2]}"
            )
        except Exception as e:
            messagebox.showerror("Eroare", str(e))

    # Butoane
    frm_btns = tk.Frame(root)
    frm_btns.pack(padx=10, pady=10, fill="x")
    tk.Button(frm_btns, text="Răsfoiește", width=14, command=choose_pdf).pack(side="left")
    tk.Button(frm_btns, text="Procesează", width=14, command=do_process).pack(side="left", padx=8)
    tk.Button(frm_btns, text="Ieșire", width=10, command=root.destroy).pack(side="right")

    root.mainloop()

if __name__ == "__main__":
    run_gui()
