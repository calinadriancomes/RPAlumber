
import os
import re
import pdfplumber
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

NUMERIC_PARCEL = re.compile(r"([1-9][0-9]{0,3}|1000)[a-i]$", re.IGNORECASE)
KEY_TOKEN      = re.compile(r"(?!(ZN|ZIE))\b(TA|TAn|CP|PP|CPn|PPn|PPi|TAi|TS)\b", re.IGNORECASE)
DECIMAL        = re.compile(r"^\d+(?:[.,]\d+)?$")
COD_PACHET     = re.compile(r"^(nu|[ABC]\d{2})$", re.IGNORECASE)

def _norm(v):
    if v is None: return ""
    s = str(v).strip()
    return re.sub(r"\s+", " ", s)

def _to_float(s):
    try:
        return float(str(s).replace(",", "."))
    except Exception:
        return None

def _extract_tables_to_excel(pdf_path, out_xlsx):
    import pdfplumber, pandas as pd
    with pdfplumber.open(pdf_path) as pdf, pd.ExcelWriter(out_xlsx) as writer:
        found_any = False
        for i, page in enumerate(pdf.pages):
            table = page.extract_table()
            if not table:
                continue
            df = pd.DataFrame(table)
            empty = pd.DataFrame([[None] * len(df.columns)], columns=df.columns)
            df = pd.concat([empty, df], ignore_index=True)
            df.to_excel(writer, sheet_name=f"Table{i+1}", index=False, header=False)
            found_any = True
        if not found_any:
            pd.DataFrame([[]]).to_excel(writer, sheet_name="Table1", index=False, header=False)

def _scan_tables_for_rows(in_xlsx):
    """Return list of parsed rows with keys: {'categorie','cultura','pachet','siruta','area'}"""
    rows = []
    import pandas as pd
    xlsx = pd.ExcelFile(in_xlsx)
    for sheet in xlsx.sheet_names:
        try:
            df = xlsx.parse(sheet, header=None)
        except Exception:
            continue
        if df is None or df.empty:
            continue
        for _, r in df.iterrows():
            vals = [_norm(v) for v in r.values.tolist()]
            if all(v == "" for v in vals):
                continue
            parcel_idx = -1
            for idx, v in enumerate(vals):
                if NUMERIC_PARCEL.search(v):
                    parcel_idx = idx
                    break
            if parcel_idx == -1:
                continue
            cat_idx = -1
            categorie = None
            for idx, v in enumerate(vals):
                if KEY_TOKEN.fullmatch(v):
                    cat_idx = idx
                    categorie = v.upper()
                    break
            if categorie is None:
                continue
            area = None
            area_idx = -1
            for idx in range(len(vals)-1, -1, -1):
                if DECIMAL.fullmatch(vals[idx]):
                    f = _to_float(vals[idx])
                    if f is not None:
                        area = f
                        area_idx = idx
                        break
            if area is None:
                continue
            cultura = None
            if parcel_idx + 1 < len(vals) and vals[parcel_idx+1] not in ("", None):
                cultura = vals[parcel_idx+1]
            if not cultura:
                span = vals[parcel_idx+1: area_idx if area_idx>parcel_idx else len(vals)]
                alpha_like = [x for x in span if re.search(r"[A-Za-zĂÂÎȘȚăâîșț]", x)]
                if alpha_like:
                    cultura = max(alpha_like, key=len)
            if not cultura:
                cand = [v for v in vals if v and not KEY_TOKEN.fullmatch(v) and not DECIMAL.fullmatch(v)]
                cultura = cand[0] if cand else "NECUNOSCUT"
            pachet = None
            siruta = None
            for v in vals:
                if COD_PACHET.fullmatch(v):
                    pachet = v.upper()
                if re.fullmatch(r"\d{6}", v):
                    siruta = v
            rows.append({
                "categorie": categorie,
                "cultura": cultura,
                "pachet": pachet,
                "siruta": siruta,
                "area": round(area, 2)
            })
    return rows

def _aggregate(rows):
    cultura_area = {}
    categorie_area = {}
    for row in rows:
        a = float(row["area"] or 0.0)
        cultura = row.get("cultura")
        categorie = row.get("categorie")
        if cultura:
            cultura_area[cultura] = round(cultura_area.get(cultura, 0.0) + a, 2)
        if categorie:
            categorie_area[categorie] = round(categorie_area.get(categorie, 0.0) + a, 2)
    return cultura_area, categorie_area

def _write_plan(plan_path, cultura_area, categorie_area, source_pdf):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Plan_culturi"
    ws1.append(["Cultura", "Suprafata_ha"])
    if cultura_area:
        for k, v in sorted(cultura_area.items()):
            ws1.append([k, v])
    for c in range(1, 3):
        ws1.column_dimensions[get_column_letter(c)].width = 28
    ws2 = wb.create_sheet("Plan_categorii")
    ws2.append(["Categorie", "Suprafata_ha"])
    if categorie_area:
        for k, v in sorted(categorie_area.items()):
            ws2.append([k, v])
    for c in range(1, 3):
        ws2.column_dimensions[get_column_letter(c)].width = 28
    meta = wb.create_sheet("Meta")
    meta.append(["Generat la", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    meta.append(["Sursa PDF", source_pdf])
    meta.append(["Observatie", "Seturile goale indică faptul că parserul nu a putut extrage rânduri valide din PDF."])
    wb.save(plan_path)

def process_file(pdf_path):
    pdf_path = os.path.abspath(pdf_path)
    out_dir = os.path.dirname(pdf_path)
    out_tables = os.path.join(out_dir, "output_tables_with_empty_row.xlsx")
    plan_path = os.path.join(out_dir, "Plan de fertilizare.xlsx")
    _extract_tables_to_excel(pdf_path, out_tables)
    rows = _scan_tables_for_rows(out_tables)
    cultura_area, categorie_area = _aggregate(rows)
    _write_plan(plan_path, cultura_area, categorie_area, pdf_path)
    return out_tables, plan_path, (len(rows), len(cultura_area), len(categorie_area))
