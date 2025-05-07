import tkinter 
import pandas as pd
import pdfplumber
from tkinter import messagebox
from tkinter import filedialog
import openpyxl
from openpyxl import workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook  
import re
from  openpyxl import *
from openpyxl.utils import get_column_letter
import xlwings as xw
import os
import pandas as pd

COORDCOLUMNparc=0
COORDCOLUMNcategorie=0
COORDCOLUMNterulet=0   
COORDCOLUMNsiruta=0
COORDCOLUMNCODpachet=0
DENUMIREcategorie=[]
DENUMIREcultura=[]
DENUMIRECODpachet=[]
CODpachet=[]
SUPRAFATAcultura=[]
SUPRAFATAcategorie=[]
SUPRAFATAcatPachet=[]

main_win = tkinter.Tk()

text0 = tkinter.Label(main_win, text = "CERERE UNICĂ DE PLATĂ din anul ",fg='black')
text0.pack()
an=tkinter.Entry(main_win)
an.insert(0,'2022')
an.pack()

text1 = tkinter.Label(main_win, text = "Selectati un document PDF'",fg='red')
text1.pack()



Shetsdecl=[]  # numar sheet care contine decl de suprafata 
ShetAdresa=[]  # nr sheet care contine date pers
ShetsdeclAnimale=[]  # numar sheet care contine date privind animalele detinute 
nrsheetgasit=0  #numarator sheet gasit decl suprafata
Shetnames=[] #oldalak nevei


def chooseFile():               # Rasfoieste file
    global location, Shetsdecl,nrsheetgasit, Shetnames 
    location=''
    directory= os.path.dirname(location)
    a=0.0   
    ta=0.0
   

    main_win.sourceFile = filedialog.askopenfilename(parent=main_win, initialdir=f"dirctory", title='Please select a directory')
    location = main_win.sourceFile
    print( str(directory), location)
               
            
   
    
    filename= str("output_tables_with_empty_row.xlsx")
    with pdfplumber.open(location) as pdf:
    # Létrehozunk egy ExcelWriter objektumot az Excel fájl írásához
        with pd.ExcelWriter(filename) as writer:
        # Végigmegyünk az összes oldalon
            for i, page in enumerate(pdf.pages):
            # Kinyerjük a táblázatot az oldalról
                table = page.extract_table()

            # Ha van táblázat az oldalon, akkor konvertáljuk DataFrame-be és mentjük Excel-be
                if table:
                # Táblázat konvertálása Pandas DataFrame-be, a teljes táblázatot adatként kezeljük
                    df = pd.DataFrame(table)

                # Létrehozunk egy üres sort az oszlopok számának megfelelően
                    empty_row = pd.DataFrame([[None] * len(df.columns)], columns=df.columns)

                # Az üres sort és a táblázatot összefűzzük
                    df = pd.concat([empty_row, df], ignore_index=True)

                # A lap neve lesz pl. "Page_1", "Page_2", stb.
                    sheet_name = f"Table{i+1}"
                    Shetnames.append(sheet_name) 

                # Az adatok exportálása Excel-be, a fejlécet kikapcsoljuk (header=False)
                    df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

                    print(f"Táblázat az {i+1}. oldalról sikeresen hozzáadva a laphoz: {sheet_name}")
                else:
                    print(f"Az {i+1}. oldalon nem található táblázat.")
                
    print(Shetnames)
    # Excel fájl elérési útja
    excel_file = "output_tables_with_empty_row.xlsx"  
   


    # Kulcsszavak és minták meghatározása
    numeric_pattern = r"([1-9][0-9]{0,2}|1000)[abcd]"  # Számok és betűk (pl. 10a, 500b)
    key_pattern = r"(?!(ZN|ZIE))\b(TA|TAn|CP|PP|CPn|PPn|PPi|TAi|TS|   )"  # Egyéb kulcsszavak
    pattern_numeric = re.compile(rf"\b{numeric_pattern}\b")  # Szám-minta
    pattern_key = re.compile(rf"\b{key_pattern}\b")  # Kulcsszó-minta

    # Talált sorok gyűjtése
    data_results = []

    try:
        # Excel fájl betöltése
        xlsx = pd.ExcelFile(excel_file)
    except FileNotFoundError:
        print(f"Hiba: Az '{excel_file}' fájl nem található.")
        exit()

    # Munkalapok feldolgozása
    for sheet_name in xlsx.sheet_names:
        print(f"Munkalap neve: {sheet_name}")
    
        try:
            # Munkalap betöltése DataFrame-be
            df = xlsx.parse(sheet_name)
        except ValueError:
            print(f"A(z) {sheet_name} munkalap üres vagy nem olvasható.")
            continue

        # Sorok iterálása
        for row_index, row in df.iterrows():
            numeric_found = False
            key_found = False

            # Cellák iterálása az adott sorban
            for value in row:
                if pd.notna(value):
                    cell_value = str(value)
                    # Keresés a szám+betű mintára
                    if pattern_numeric.search(cell_value):
                        numeric_found = True
                    # Keresés a kulcsszóra
                    if pattern_key.fullmatch(cell_value):
                        key_found = True

            # Ha mindkettő megtalálható, hozzáadjuk az eredményeket
            if numeric_found and key_found:
                data_results.append(row.values)  # Sor értékei listaként

    # Eredmények mentése ugyanabba az Excel fájlba
    if data_results:
        # Eredmények DataFrame-ként
        results_df = pd.DataFrame(data_results)
    
        # Excel fájl megnyitása íráshoz (openpyxl)
        with pd.ExcelWriter(excel_file, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            # Eredményeket DATA nevű munkalapra mentjük
            results_df.to_excel(writer, sheet_name="DATA", index=False, header=False)
    
        print(f"Talált sorok elmentve a(z) '{excel_file}' fájlban a 'DATA' munkalapra.")
    else:
        print("Nem található olyan sor, amely megfelelne a feltételeknek.")


        # Megnyitjuk az excel file-table
    wb = openpyxl.open("output_tables_with_empty_row.xlsx") 
        #create new exxel sheet

    if 'DATAA' not in wb.sheetnames:
        wb.create_sheet('DATAA') 
    
    filename_out=f"{location} output_tables_with_empty_row.xlsx"  
    print(filename_out)
    wb.save(filename_out)
    
    
    
    
    
    filename_pf=f"{location} Plan de fertilizare.xlsx"
    # Ellenőrzés: létezik-e a fájl az aktuális mappában
    if not os.path.exists(filename_pf):
    
        pf= xw.Book()
        pf.save(filename_pf)
        print(f"A '{filename_pf}' fájl letre lett hozva")
    else:
        print(f"A '{filename_pf}' fájl már létezik.")
    
    # Excel fájl és munkalap megnyitása
    wb = xw.Book(filename_out)  # Excel fájl betöltése
              
    sheet = wb.sheets["DATA"]  # Munkalap kiválasztása

    #    Adatok beolvasása (használati tartomány)
    data = sheet.used_range.value  # Ez egy listák listája (Excel tartalom)

    # Regex minta definiálása
    decimal_pattern = r"^\d+(\.\d+)$"  # Tizedes számok (pl. 12.34, 100.0)
    numeric_pattern = r"([1-9][0-9]{0,3}|1000)[abcdefghi]"  # Számok és betűk (pl. 10a, 500b)
    #key_pattern = r"(?!(ZN|ZIE))\b(TA|TAn|CP|PP|CPn|PPn|PPi|TAi|TS|   )"  # Egyéb kulcsszavak
    key_siruta=r"([0-9]{6})"
    key_codpachet=r"(nu|B[0-9][0-9]|A[0-9][0-9]|C[0-9][0-9])"

    used_range = sheet.used_range   # Az első nem üres oszlop
    first_col = used_range.column
    first_row = used_range.row       
    last_row = used_range.last_cell.row # Az utolsó nem üres sor
    last_col = used_range.last_cell.column # Az utolsó nem üres oszlop

    # Ellenőrzés és adatfeldolgozás
    for row in range(1, 2 ):  # 1-től az utolsó sorig   
        for col in range(1, last_col+1 ):  # 1-től az utolsó oszlopig
            cell_value = sheet.cells(row, col).value
            if isinstance(cell_value, str) and re.match(decimal_pattern, cell_value):
                if col==last_col:
                    COORDCOLUMNterulet=int(col)
                    print(f"{row},{col}: Tizedes számot találtunk: {cell_value}")
                    print(f"COORDCOLUMNterulet: {COORDCOLUMNterulet}")
            elif isinstance(cell_value, str) and re.match(numeric_pattern, cell_value):        
                COORDCOLUMNparc=int(col)  
                print(f"{row},{col}: Parcella számot találtunk: {cell_value}")
                print(f"COORDCOLUMNparc: {COORDCOLUMNparc}")
            elif isinstance(cell_value, str) and re.match(key_pattern, cell_value):        
                COORDCOLUMNcategorie=int(col)  
                print(f"{row},{col}: Kategoria jelzest találtunk: {cell_value}")        
                print(f"COORDCOLUMNcategorie: {COORDCOLUMNcategorie}")    
            elif isinstance(cell_value, str) and re.match(key_siruta, cell_value):        
                COORDCOLUMNsiruta=int(col)
                print(f"{row},{col}: Cod siruta jelzest találtunk: {cell_value}")        
                print(f"COORDCOLUMNsiruta: {COORDCOLUMNsiruta}")   
            elif isinstance(cell_value, str) and re.match(key_codpachet, cell_value):        
                COORDCOLUMNCODpachet=int(col)
                print(f"{row},{col}: Cod Pachet jelzest találtunk: {cell_value}")        
                print(f"COORDCOLUMNCODpachet {COORDCOLUMNCODpachet}")                     

    print(last_row)  
              
    for q in range(1, int(last_row)+1 ):  # 1-től az utolsó sorig
        cell_value = sheet.cells(q, COORDCOLUMNcategorie).value  
    
        if cell_value is not None and cell_value not in DENUMIREcategorie:
            DENUMIREcategorie.append(cell_value)
       
        print(q,cell_value, DENUMIREcategorie)
    
    
    print(last_row) 
    for p in range(1, int(last_row)+1 ):  # 1-től az utolsó sorig
        cell_value = sheet.cells(p, COORDCOLUMNparc+1).value  
        if cell_value is not None and cell_value not in DENUMIREcultura:
            DENUMIREcultura.append(cell_value)
            print(f"{p}:{row},{col},{cell_value} ") 

    for c in range(1, int(last_row)+1 ):  # 1-től az utolsó sorig
        cell_value = sheet.cells(c, COORDCOLUMNCODpachet).value  
        if cell_value is not None and cell_value not in DENUMIRECODpachet:
            DENUMIRECODpachet.append(cell_value)
            print(f"{c}:{row},{col},{cell_value} ")
            
    LENGTH_DENUMIRECODpachet=len(DENUMIRECODpachet)
    LENGTH_DENUMIREcultura=len(DENUMIREcultura)
    LENGTH_DENUMIREcategorie=len(DENUMIREcategorie)

    for i in range(0,LENGTH_DENUMIREcultura): 
        cultura=DENUMIREcultura[i]
        SUPRAFATAcultura.append(0.00)
    
        for q in range(1, int(last_row)+1 ):  # 1-től az utolsó sorig
            cell_value = sheet.cells(q, COORDCOLUMNparc+1).value 
            areacell_value=sheet.cells(q, COORDCOLUMNterulet).value
            if cell_value is not None and cultura==cell_value:    
            
                SUPRAFATAcultura[i]=round(SUPRAFATAcultura[i] + float(areacell_value), 2)
                print(cell_value,float(areacell_value), SUPRAFATAcultura[i]   )    
        sheet.cells(last_row+3+i,last_col-4).value =SUPRAFATAcultura[i]
        sheet.cells(last_row+3+i,last_col-5).value =cultura
        print(DENUMIREcultura, SUPRAFATAcultura    )   

    for i in range(0,LENGTH_DENUMIREcategorie): 
    
        categorie=DENUMIREcategorie[i]
        SUPRAFATAcategorie.append(0.00)
        for q in range(1, int(last_row)+1 ):  # 1-től az utolsó sorig
            cell_value = sheet.cells(q, COORDCOLUMNcategorie).value 
            areacell_value=sheet.cells(q, COORDCOLUMNterulet).value
           
            if cell_value is not None:    
                if categorie==cell_value:
                    SUPRAFATAcategorie[i]=round(SUPRAFATAcategorie[i] + float(areacell_value), 2)
                    print(cell_value,float(areacell_value), SUPRAFATAcategorie[i]   )    
        sheet.cells(last_row+3+i,last_col-7).value =SUPRAFATAcategorie[i]
        sheet.cells(last_row+3+i,last_col-8).value =categorie
        print(DENUMIREcategorie, SUPRAFATAcategorie    )        

    for i in range(0,LENGTH_DENUMIRECODpachet): 
    
        pachet=DENUMIRECODpachet[i]
        SUPRAFATAcatPachet.append(0.00)
        for q in range(1, int(last_row)+1 ):  # 1-től az utolsó sorig
            cell_value = sheet.cells(q, COORDCOLUMNCODpachet).value 
            areacell_value=sheet.cells(q, COORDCOLUMNterulet).value
           
            if cell_value is not None:    
                if pachet==cell_value:
                    SUPRAFATAcatPachet[i]=round(SUPRAFATAcatPachet[i] + float(areacell_value), 2)
                    print(cell_value,float(areacell_value), SUPRAFATAcatPachet[i]   )    
        sheet.cells(last_row+3+i,last_col).value =SUPRAFATAcatPachet[i]
        sheet.cells(last_row+3+i,last_col-1).value =pachet
        print(DENUMIRECODpachet, SUPRAFATAcatPachet    ) 

    TOTAL=0.00
    for i in range(0,len(SUPRAFATAcategorie)):
        Supr=float(SUPRAFATAcategorie[i])
   
        TOTAL +=Supr  
    sheet.cells(last_row+2+i,last_col-10).value =TOTAL
    sheet.cells(last_row+2+i,last_col-11).value ="TOTAL"
         
    try:
        wb.save(filename_out)
    except Exception as e:
        print(f"Hiba mentés közben: {e}")
    finally:
        wb.close()            

   # wb.save()
   # wb = xw.Book(filename_out)
   
b_chooseFile = tkinter.Button(main_win, text = "Răsfoiește", width = 10, height = 1, command = chooseFile)
b_chooseFile.pack()
b_chooseTab = tkinter.Button(main_win, text = "Exportă datele", width = 10, height = 1)
b_chooseTab.pack()
b_chooseTab = tkinter.Button(main_win, text = "Calcul", width = 10, height = 1)
b_chooseTab.pack()
b_exit = tkinter.Button(main_win, text = "Exit",command = main_win.destroy)
b_exit.pack()

main_win.mainloop()
